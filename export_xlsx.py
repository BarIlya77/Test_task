import os

import openpyxl

headers = ["Temperature", "Precipitation", "Pressure", "Wind Speed", "Wind Direction"]
sheet_title = "Weather Data"
file_name = "weather_data.xlsx"

def write_data_xlsx(data):
    filename = file_name
    if os.path.exists(filename):
        workbook = openpyxl.load_workbook(filename)
        index = len(workbook.sheetnames)
        sheet = workbook.create_sheet(title=f"{sheet_title} {index}")
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_title

    sheet.append(headers)
    for item in data:
        sheet.append(item)

    workbook.save("weather_data.xlsx")
    return "Данные успешно записаны в файл weather_data.xlsx"
